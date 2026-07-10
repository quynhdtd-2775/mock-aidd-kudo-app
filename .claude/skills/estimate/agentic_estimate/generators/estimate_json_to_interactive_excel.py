"""Render estimate JSON to interactive Excel workbook with live formulas."""

from __future__ import annotations

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter as gcl
except ImportError:
    Workbook = None

from agentic_estimate.generators.estimate_render_helpers import (
    all_tasks_from_option,
    get_roles,
    role_display,
    task_role_md,
    task_total_md,
)

BLUE, L_BLUE, GREEN, L_GREEN = "4472C4", "D9E2F3", "70AD47", "E2EFDA"
ORANGE, L_ORANGE, L_YELLOW, WHITE, DARK = "ED7D31", "FCE4D6", "FFF2CC", "FFFFFF", "2F2F2F"
THIN = None
ROLE_COLORS = [
    ("4472C4", "D9E2F3"),
    ("70AD47", "E2EFDA"),
    ("ED7D31", "FCE4D6"),
    ("7030A0", "E2D0F0"),
    ("C00000", "F4CCCC"),
]


def _init_border():
    global THIN
    if THIN is not None:
        return
    THIN = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def _hdr(bg=BLUE):
    return {
        "font": Font(bold=True, color=WHITE, size=11),
        "fill": PatternFill(start_color=bg, end_color=bg, fill_type="solid"),
        "border": THIN,
        "alignment": Alignment(horizontal="center", wrap_text=True),
    }


def _sty(bg=None):
    s = {"border": THIN, "alignment": Alignment(vertical="center")}
    if bg:
        s["fill"] = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    return s


def _apply(c, s):
    for k, v in s.items():
        setattr(c, k, v)


def _c(ws, r, c, v, s=None, f=None, b=False):
    """Set cell: value, style, format, bold."""
    cell = ws.cell(row=r, column=c, value=v)
    if s:
        _apply(cell, s)
    if f:
        cell.number_format = f
    if b:
        cell.font = (
            Font(bold=True) if not hasattr(cell, "font") else Font(bold=True, color=cell.font.color)
        )
    return cell


def _rc(i):
    return ROLE_COLORS[i % len(ROLE_COLORS)]


def _rmd(row, role, cm):
    b, c, e, buf = [gcl(cm[f"{role}_{k}"]) for k in ["base", "cpx", "exp", "buf"]]
    return f"=ROUND({b}{row}*{c}{row}*{e}{row}*(1+{buf}{row}),2)"


def _tmd(row, roles, cm):
    refs = [f"{gcl(cm[f'{r}_md'])}{row}" for r in roles]
    return f"={'+'.join(refs)}" if refs else "=0"


def _add_param(ws, row, key, val, desc, cells):
    """Add parameter row and return next row number."""
    _c(ws, row, 1, key, _sty(L_BLUE), b=True)
    _c(ws, row, 2, desc, _sty())
    c = _c(ws, row, 3, val, _sty(L_YELLOW), "#,##0" if "cost" in key else "0.00", True)
    c.font = Font(bold=True, size=12, color=DARK)
    _c(ws, row, 4, val, _sty())
    cells[key] = f"Parameters!$C${row}"
    return row + 1


def _params_sheet(wb, data):
    ws = wb.create_sheet("Parameters", 0)
    ws.sheet_properties.tabColor = BLUE
    ws.cell(1, 1, "PROJECT-LEVEL MULTIPLIERS").font = Font(bold=True, size=14, color=BLUE)
    ws.merge_cells("A1:D1")
    ws.cell(2, 1, "Change column C to recalculate all estimates.").font = Font(
        italic=True, color="888888"
    )
    for i, lbl in enumerate(["Parameter", "Description", "Value", "KB Default"], 1):
        _c(ws, 4, i, lbl, _hdr())
    params, cells, row = data.get("parameters", {}), {}, 5
    for m in params.get("project_multipliers", []):
        row = _add_param(ws, row, m["key"], m["value"], m.get("rationale", ""), cells)
    if cbr := params.get("cost_per_md_by_role", {}):
        roles, names = get_roles(data)
        for r in roles:
            row = _add_param(
                ws,
                row,
                f"cost_per_md_{r}",
                cbr.get(r, 40000),
                f"Cost per MD ({role_display(r, names)})",
                cells,
            )
    elif "cost_per_md" not in cells:
        row = _add_param(
            ws, row, "cost_per_md", params.get("cost_per_md", 40000), "Cost per man-day", cells
        )
    if "ai_reduction" not in cells:
        opt_a = next((o for o in data.get("options", []) if o.get("categories")), {})
        row = _add_param(
            ws,
            row,
            "ai_reduction",
            (100 - opt_a.get("ai_reduction_pct", 20)) / 100,
            "AI productivity multiplier",
            cells,
        )
    for i, w in enumerate([22, 50, 14, 14], 1):
        ws.column_dimensions[gcl(i)].width = w
    return cells


def _option_a_sheet(wb, data, pc):
    opt = next((o for o in data.get("options", []) if o.get("categories")), None)
    if not opt:
        return
    roles, names = get_roles(data)
    ws = wb.create_sheet("Option A")
    ws.sheet_properties.tabColor = GREEN
    ws.cell(1, 1, f"OPTION A — {opt.get('subtitle', 'Full Scope')}").font = Font(
        bold=True, size=14, color=GREEN
    )
    ws.cell(2, 1, "Edit role Base/Cpx/Exp/Buf% cells. MD auto-recalculates.").font = Font(
        italic=True, color="888888"
    )
    cm, col = {}, 4
    for r in roles:
        for i, k in enumerate(["base", "cpx", "exp", "buf", "md"]):
            cm[f"{r}_{k}"] = col + i
        col += 5
    for i, k in enumerate(["total_md", "sp", "with_ai", "orig_md", "delta", "notes"]):
        cm[k] = col + i
    for idx, r in enumerate(roles):
        hbg, _ = _rc(idx)
        _c(ws, 3, cm[f"{r}_base"], role_display(r, names), _hdr(hbg))
        ws.merge_cells(
            start_row=3, start_column=cm[f"{r}_base"], end_row=3, end_column=cm[f"{r}_md"]
        )
        for i, lbl in enumerate(["Base", "Cpx", "Exp", "Buf%", "MD"]):
            _c(ws, 4, cm[f"{r}_base"] + i, lbl, _hdr(hbg))
    for lbl in ["Total MD", "SP", "With AI", "Orig MD", "Delta", "Notes"]:
        _c(ws, 4, cm[lbl.lower().replace(" ", "_")], lbl, _hdr(GREEN))
    for col, lbl in enumerate(["ID", "Task", "Category"], 1):
        _c(ws, 3, col, lbl, _hdr(GREEN))
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)

    ai_cell, row, start = pc.get("ai_reduction", ""), 5, 5
    for cat in opt.get("categories", []):
        for t in cat.get("tasks", []):
            _c(ws, row, 1, t["id"], _sty())
            _c(ws, row, 2, t["name"], _sty())
            _c(ws, row, 3, cat["name"], _sty(L_BLUE))
            for idx, r in enumerate(roles):
                _, lt = _rc(idx)
                eff = t.get("effort", {}).get(r, {})
                _c(ws, row, cm[f"{r}_base"], eff.get("base", 0), _sty(lt), "0.0")
                _c(ws, row, cm[f"{r}_cpx"], eff.get("complexity", 1.0), _sty(lt), "0.00")
                _c(ws, row, cm[f"{r}_exp"], eff.get("experience", 1.0), _sty(lt), "0.00")
                _c(ws, row, cm[f"{r}_buf"], eff.get("buffer_pct", 15) / 100, _sty(lt), "0%")
                _c(ws, row, cm[f"{r}_md"], _rmd(row, r, cm), _sty(lt), b=True)
            tc = gcl(cm["total_md"])
            _c(ws, row, cm["total_md"], _tmd(row, roles, cm), _sty(L_GREEN), b=True)
            _c(ws, row, cm["sp"], t.get("story_points", 0), _sty())
            ai_f = (
                f"=ROUND({tc}{row}*{ai_cell},2)"
                if t.get("is_dev_task", False) and ai_cell
                else f"={tc}{row}"
            )
            _c(ws, row, cm["with_ai"], ai_f, _sty(L_GREEN))
            _c(ws, row, cm["orig_md"], t.get("original_md", 0), _sty())
            _c(ws, row, cm["delta"], f"={tc}{row}-{gcl(cm['orig_md'])}{row}", _sty())
            _c(ws, row, cm["notes"], (t.get("notes", "") or "")[:80], _sty())
            row += 1

    end, tr = row - 1, row
    _c(ws, row, 2, "TOTAL", _sty(L_GREEN), b=True).font = Font(bold=True, size=12)
    for k in [f"{r}_md" for r in roles] + ["total_md", "sp", "with_ai", "orig_md", "delta"]:
        _c(ws, row, cm[k], f"=SUM({gcl(cm[k])}{start}:{gcl(cm[k])}{end})", _sty(L_GREEN), b=True)
    row += 1
    _c(ws, row, 2, "COST", _sty(L_ORANGE), b=True).font = Font(bold=True, size=12)
    for r in roles:
        if cost := pc.get(f"cost_per_md_{r}", pc.get("cost_per_md", "")):
            _c(
                ws,
                row,
                cm[f"{r}_md"],
                f"={gcl(cm[f'{r}_md'])}{tr}*{cost}",
                _sty(L_ORANGE),
                "¥#,##0",
                True,
            )
    if gc := pc.get("cost_per_md", ""):
        for k in ["total_md", "with_ai"]:
            _c(ws, row, cm[k], f"={gcl(cm[k])}{tr}*{gc}", _sty(L_ORANGE), "¥#,##0", True)
    (
        ws.column_dimensions["A"].width,
        ws.column_dimensions["B"].width,
        ws.column_dimensions["C"].width,
    ) = (6, 38, 20)
    for r in roles:
        for i, w in enumerate([8, 8, 6, 7, 8]):
            ws.column_dimensions[gcl(cm[f"{r}_base"] + i)].width = w
    for k, w in [
        ("total_md", 9),
        ("sp", 6),
        ("with_ai", 9),
        ("orig_md", 9),
        ("delta", 7),
        ("notes", 30),
    ]:
        ws.column_dimensions[gcl(cm[k])].width = w
    ws.freeze_panes = "D5"


def _option_b_sheet(wb, data, pc):
    opt = next((o for o in data.get("options", []) if o.get("scope_reductions")), None)
    if not opt:
        return
    roles, names = get_roles(data)
    ws = wb.create_sheet("Option B")
    ws.sheet_properties.tabColor = ORANGE
    ws.cell(1, 1, f"OPTION B — {opt.get('subtitle', 'Budget')}").font = Font(
        bold=True, size=14, color=ORANGE
    )
    ws.cell(2, 1, "Per-role MD breakdown.").font = Font(italic=True, color="888888")

    hdrs = (
        ["ID", "Task"]
        + [role_display(r, names) for r in roles]
        + ["Total B", "SP", "A MD", "Reduction", "Cost"]
    )
    for i, h in enumerate(hdrs, 1):
        _c(ws, 4, i, h, _hdr(ORANGE))

    cc, row, start, co, tbc = pc.get("cost_per_md", ""), 5, 5, 3, 3 + len(roles)
    for t in all_tasks_from_option(opt):
        _c(ws, row, 1, t["id"], _sty())
        _c(ws, row, 2, t["name"], _sty())
        for idx, r in enumerate(roles):
            _c(ws, row, co + idx, task_role_md(t, r) or 0, _sty())
        _c(ws, row, tbc, task_total_md(t), _sty(L_YELLOW))
        _c(ws, row, tbc + 1, t.get("story_points", 0), _sty())
        _c(ws, row, tbc + 2, t.get("option_a_md", 0), _sty())
        _c(ws, row, tbc + 3, (t.get("reduction_reason", "") or "")[:50], _sty())
        if cc:
            _c(ws, row, tbc + 4, f"={gcl(tbc)}{row}*{cc}", _sty(), "¥#,##0")
        row += 1
    end, tr = row - 1, row
    _c(ws, row, 2, "TOTAL", _sty(L_ORANGE), b=True).font = Font(bold=True, size=12)
    for col in range(co, tbc + 3):
        _c(ws, row, col, f"=SUM({gcl(col)}{start}:{gcl(col)}{end})", _sty(L_ORANGE), b=True)
    if cc:
        _c(
            ws,
            row,
            tbc + 4,
            f"=SUM({gcl(tbc + 4)}{start}:{gcl(tbc + 4)}{end})",
            _sty(L_ORANGE),
            "¥#,##0",
            True,
        )
    if budget := opt.get("budget_target"):
        row += 2
        ws.cell(row, 1, "Budget Target")
        _c(ws, row, 2, budget, _sty(L_YELLOW), "¥#,##0", True)
        row += 1
        ws.cell(row, 1, "Over/Under")
        c = ws.cell(row, 2, f"={gcl(tbc + 4)}{tr}-B{row-1}")
        c.number_format, c.font = "¥#,##0", Font(bold=True)
    ws.column_dimensions["A"].width, ws.column_dimensions["B"].width = 7, 35
    for i in range(len(roles)):
        ws.column_dimensions[gcl(co + i)].width = 8
    for i, w in enumerate([9, 6, 8, 30, 14], tbc):
        ws.column_dimensions[gcl(i)].width = w
    ws.freeze_panes = "A5"


def render(data: dict, output_path: Path) -> Path:
    if not Workbook:
        raise ImportError("openpyxl required: pip install openpyxl")
    _init_border()
    wb = Workbook()
    cells = _params_sheet(wb, data)
    _option_a_sheet(wb, data, cells)
    _option_b_sheet(wb, data, cells)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(str(output_path))
    return output_path
