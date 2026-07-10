"""Excel generator using OpenPyXL."""

from __future__ import annotations

from datetime import datetime
from typing import Any

OPENPYXL_AVAILABLE = False
Workbook = None
Font = None
PatternFill = None
Alignment = None
Border = None
Side = None
get_column_letter = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: F401
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    pass

from .generator_base_output_interface import BaseGenerator, OutputConfig


class ExcelGenerator(BaseGenerator):
    """Generates Excel spreadsheets for estimates."""

    def __init__(self, config: OutputConfig | None = None):
        """Initialize generator with styles."""
        super().__init__(config)
        self._init_styles()

    def _init_styles(self) -> None:
        """Initialize styles if openpyxl is available."""
        if OPENPYXL_AVAILABLE:
            self.HEADER_FILL = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            self.HEADER_FONT = Font(bold=True, color="FFFFFF")
            self.SECTION_FILL = PatternFill(
                start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
            )
            self.THIN_BORDER = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
        else:
            self.HEADER_FILL = None
            self.HEADER_FONT = None
            self.SECTION_FILL = None
            self.THIN_BORDER = None

    @property
    def file_extension(self) -> str:
        return "xlsx"

    def generate(self, results: dict) -> str:
        """Generate Excel spreadsheet."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel generation. Install with: pip install openpyxl"
            )

        wb = Workbook()

        self._create_summary_sheet(wb, results)
        self._create_requirements_sheet(wb, results)
        self._create_phases_sheet(wb, results)

        if self.config.include_risks:
            self._create_risks_sheet(wb, results)

        if self.config.include_api_spec:
            self._create_architecture_sheet(wb, results)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        filename = self.get_filename()
        output_path = self.config.get_output_path(filename)
        wb.save(output_path)

        return str(output_path)

    def _create_summary_sheet(self, wb: Any, results: dict) -> None:
        """Create summary sheet."""
        ws = wb.create_sheet("Summary", 0)
        estimate = results.get("estimate", {})
        validation = results.get("validation_report", {})
        risk_assessment = results.get("risk_assessment", {})

        ws["A1"] = f"Project Estimate: {results.get('project_name', 'Project')}"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")

        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A3"] = ""

        # Summary metrics
        summary_data = [
            ["Metric", "Value", "Notes"],
            ["Total Story Points", estimate.get("total_story_points", 0), ""],
            ["Total Man-Days", estimate.get("total_man_days", 0), ""],
            ["Buffer", f"{estimate.get('buffer_percentage', 20)}%", ""],
            ["Confidence Level", estimate.get("confidence_level", "medium"), ""],
            ["Risk Level", risk_assessment.get("overall_level", "medium"), ""],
            ["Quality Score", f"{validation.get('quality_score', 'N/A')}/100", ""],
            ["Validation Status", validation.get("status", "N/A"), ""],
        ]

        start_row = 4
        for i, row in enumerate(summary_data):
            for j, value in enumerate(row):
                cell = ws.cell(row=start_row + i, column=j + 1, value=value)
                if i == 0:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT
                cell.border = self.THIN_BORDER

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 30

    def _create_requirements_sheet(self, wb: Any, results: dict) -> None:
        """Create requirements breakdown sheet."""
        ws = wb.create_sheet("Requirements")
        requirements = results.get("requirements", [])

        # Header
        headers = ["ID", "Title", "Type", "Priority", "Complexity", "Story Points", "Man-Days"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER

        # Data
        for row, req in enumerate(requirements, 2):
            ws.cell(row=row, column=1, value=req.get("id", "")).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=req.get("title", "")[:100]).border = self.THIN_BORDER
            ws.cell(row=row, column=3, value=req.get("type", "")).border = self.THIN_BORDER
            ws.cell(row=row, column=4, value=req.get("priority", "")).border = self.THIN_BORDER
            ws.cell(row=row, column=5, value=req.get("complexity", "")).border = self.THIN_BORDER
            ws.cell(row=row, column=6, value=req.get("story_points", 0)).border = self.THIN_BORDER
            ws.cell(row=row, column=7, value=req.get("man_days", 0)).border = self.THIN_BORDER

        # Totals row
        total_row = len(requirements) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{total_row-1})")
        ws.cell(row=total_row, column=7, value=f"=SUM(G2:G{total_row-1})")

        # Column widths
        widths = [10, 50, 15, 12, 12, 15, 12]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _create_phases_sheet(self, wb: Any, results: dict) -> None:
        """Create phases breakdown sheet."""
        ws = wb.create_sheet("Phases")
        estimate = results.get("estimate", {})
        phases = estimate.get("phases", [])

        # Header
        headers = ["Phase", "Story Points", "Man-Days", "% of Total", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER

        total_sp = estimate.get("total_story_points", 1)

        # Data
        for row, phase in enumerate(phases, 2):
            sp = phase.get("story_points", 0)
            pct = round((sp / total_sp) * 100, 1) if total_sp > 0 else 0

            ws.cell(row=row, column=1, value=phase.get("name", "")).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=sp).border = self.THIN_BORDER
            ws.cell(row=row, column=3, value=phase.get("man_days", 0)).border = self.THIN_BORDER
            ws.cell(row=row, column=4, value=f"{pct}%").border = self.THIN_BORDER
            ws.cell(row=row, column=5, value=phase.get("description", "")).border = self.THIN_BORDER

        # Column widths
        widths = [25, 15, 12, 12, 50]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _create_risks_sheet(self, wb: Any, results: dict) -> None:
        """Create risk assessment sheet."""
        ws = wb.create_sheet("Risks")
        risk_assessment = results.get("risk_assessment", {})
        risks = risk_assessment.get("risks", [])

        # Summary section
        ws["A1"] = "Risk Assessment Summary"
        ws["A1"].font = Font(bold=True, size=14)

        ws["A3"] = "Overall Level:"
        ws["B3"] = risk_assessment.get("overall_level", "medium")
        ws["A4"] = "Total Score:"
        ws["B4"] = risk_assessment.get("total_score", 0)

        # Risk table
        headers = ["Category", "Type", "Impact", "Score", "Level", "Mitigation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER

        for row, risk in enumerate(risks, 7):
            mitigations = risk.get("mitigations", [])
            mitigation_text = mitigations[0] if mitigations else ""

            ws.cell(row=row, column=1, value=risk.get("category", "")).border = self.THIN_BORDER
            ws.cell(
                row=row, column=2, value=risk.get("type", "").replace("_", " ").title()
            ).border = self.THIN_BORDER
            ws.cell(row=row, column=3, value=risk.get("impact", "")).border = self.THIN_BORDER
            ws.cell(row=row, column=4, value=risk.get("score", 0)).border = self.THIN_BORDER
            ws.cell(row=row, column=5, value=risk.get("level", "")).border = self.THIN_BORDER
            ws.cell(row=row, column=6, value=mitigation_text).border = self.THIN_BORDER

        # Column widths
        widths = [18, 25, 10, 10, 10, 50]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _create_architecture_sheet(self, wb: Any, results: dict) -> None:
        """Create architecture overview sheet."""
        ws = wb.create_sheet("Architecture")
        architecture = results.get("architecture", {})

        ws["A1"] = "Architecture Overview"
        ws["A1"].font = Font(bold=True, size=14)

        row = 3

        # Tech Stack
        tech_stack = architecture.get("tech_stack", {})
        if tech_stack:
            ws.cell(row=row, column=1, value="Tech Stack").font = Font(bold=True)
            row += 1

            for category, items in tech_stack.items():
                ws.cell(row=row, column=1, value=category.title())
                if isinstance(items, list):
                    ws.cell(row=row, column=2, value=", ".join(items))
                else:
                    ws.cell(row=row, column=2, value=str(items))
                row += 1

            row += 1

        # Components
        components = architecture.get("components", [])
        if components:
            ws.cell(row=row, column=1, value="Components").font = Font(bold=True)
            row += 1

            for comp in components[:15]:
                ws.cell(row=row, column=1, value=comp.get("name", ""))
                ws.cell(row=row, column=2, value=comp.get("description", ""))
                row += 1

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 60
