"""Q&A log JSON to categorized Excel workbook generator."""

import json
from pathlib import Path

import jsonschema
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

SCHEMA_PATH = Path(__file__).parents[2] / "schemas" / "qa-log-schema.json"

CATEGORY_COLORS = {
    "tech": "4472C4",
    "business": "70AD47",
    "infra": "ED7D31",
}

STATUS_COLORS = {
    "open": "FFC7CE",
    "answered": "C6EFCE",
}


def _validate_data(data: dict) -> None:
    """Validate Q&A log data against JSON schema."""
    with open(SCHEMA_PATH, encoding="utf-8") as f:
        schema = json.load(f)
    jsonschema.validate(instance=data, schema=schema)


def _apply_header_style(ws: Worksheet, category: str) -> None:
    """Apply styling to header row."""
    fill_color = CATEGORY_COLORS.get(category, "808080")
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font


def _apply_status_conditional_formatting(ws: Worksheet, row_count: int) -> None:
    """Apply conditional formatting based on status column."""
    for row in range(2, row_count + 1):
        status_cell = ws[f"D{row}"]
        status_value = status_cell.value

        if status_value == "open":
            status_cell.fill = PatternFill(
                start_color=STATUS_COLORS["open"],
                end_color=STATUS_COLORS["open"],
                fill_type="solid",
            )
        elif status_value == "answered":
            status_cell.fill = PatternFill(
                start_color=STATUS_COLORS["answered"],
                end_color=STATUS_COLORS["answered"],
                fill_type="solid",
            )


def _auto_adjust_columns(ws: Worksheet, max_width: int = 60) -> None:
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        column = col[0].column_letter
        max_length = 0

        for cell in col:
            try:
                cell_value = str(cell.value) if cell.value else ""
                max_length = max(max_length, len(cell_value))
            except Exception:
                pass

        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column].width = adjusted_width


def _create_sheet(
    wb: Workbook, sheet_name: str, questions: list[dict], category: str | None = None
) -> None:
    """Create a worksheet with questions data."""
    ws = wb.create_sheet(sheet_name)

    headers = ["ID", "Question", "Answer", "Status", "Priority", "Source", "Notes"]
    ws.append(headers)

    filtered_questions = questions
    if category:
        filtered_questions = [q for q in questions if q.get("category") == category]

    for q in filtered_questions:
        ws.append(
            [
                q.get("id", ""),
                q.get("question", ""),
                q.get("answer", ""),
                q.get("status", "open"),
                q.get("priority", "medium"),
                q.get("source", ""),
                q.get("notes", ""),
            ]
        )

    _apply_header_style(ws, category or "summary")
    _apply_status_conditional_formatting(ws, len(filtered_questions) + 1)
    _auto_adjust_columns(ws)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def render(data: dict, output_path: str | Path, config: dict | None = None) -> None:
    """Render Q&A log JSON to categorized XLSX.

    Args:
        data: Q&A log data matching qa-log-schema.json
        output_path: Path to write Excel file
        config: Optional rendering configuration (unused)

    Raises:
        jsonschema.ValidationError: If data doesn't match schema
    """
    _validate_data(data)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    questions = data.get("questions", [])

    _create_sheet(wb, "Summary", questions)
    _create_sheet(wb, "Tech", questions, "tech")
    _create_sheet(wb, "Business", questions, "business")
    _create_sheet(wb, "Infra", questions, "infra")

    wb.save(output_path)
