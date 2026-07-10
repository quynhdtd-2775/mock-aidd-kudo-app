"""Sheet builders for standalone WBS XLSX generator."""

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def create_summary_sheet(wb, wbs_data: list, active_roles: list, role_names: dict) -> None:
    """Create WBS Summary sheet with all tasks."""
    ws = wb.create_sheet("WBS Summary")

    headers = ["ID", "Category", "Task"]
    for role in active_roles:
        role_name = role_names.get(role, role.upper())
        headers.append(f"{role_name} MD")
    headers.extend(["Total MD", "Story Points", "Notes"])

    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    current_category = None
    category_start_row = 2
    row_num = 2

    for task in wbs_data:
        category = task["category"]

        if current_category != category:
            if current_category is not None:
                _add_category_subtotal(
                    ws, row_num, category_start_row, active_roles, current_category
                )
                row_num += 1

            current_category = category
            category_start_row = row_num

        row_data = [task["id"], category, task["name"]]

        for role in active_roles:
            md = task["effort"].get(role, {}).get("md", 0) or 0
            row_data.append(md if md > 0 else "")

        row_data.extend([task["total_md"], task["story_points"], task["notes"]])

        ws.append(row_data)
        row_num += 1

    if current_category is not None:
        _add_category_subtotal(ws, row_num, category_start_row, active_roles, current_category)
        row_num += 1

    _add_grand_total(ws, row_num, active_roles)

    ws.freeze_panes = "A2"
    auto_adjust_columns(ws)


def _add_category_subtotal(
    ws, row_num: int, start_row: int, active_roles: list, category: str
) -> None:
    """Add subtotal row for category."""
    subtotal_row = ["", f"{category} Subtotal", ""]

    col_offset = 3
    for idx in range(len(active_roles)):
        col_letter = get_column_letter(col_offset + idx + 1)
        formula = f"=SUM({col_letter}{start_row}:{col_letter}{row_num - 1})"
        subtotal_row.append(formula)

    total_md_col = get_column_letter(col_offset + len(active_roles) + 1)
    subtotal_row.append(f"=SUM({total_md_col}{start_row}:{total_md_col}{row_num - 1})")
    subtotal_row.extend(["", ""])

    ws.append(subtotal_row)

    for cell in ws[row_num]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")


def _add_grand_total(ws, row_num: int, active_roles: list) -> None:
    """Add grand total row at bottom."""
    total_row = ["", "Grand Total", ""]

    col_offset = 3
    for idx in range(len(active_roles)):
        col_letter = get_column_letter(col_offset + idx + 1)
        formula = f"=SUBTOTAL(9,{col_letter}2:{col_letter}{row_num - 1})"
        total_row.append(formula)

    total_md_col = get_column_letter(col_offset + len(active_roles) + 1)
    total_row.append(f"=SUBTOTAL(9,{total_md_col}2:{total_md_col}{row_num - 1})")
    total_row.extend(["", ""])

    ws.append(total_row)

    for cell in ws[row_num]:
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")


def create_per_role_sheets(wb, wbs_data: list, active_roles: list, role_names: dict) -> None:
    """Create per-role sheets with filtered tasks."""
    for role in active_roles:
        role_name = role_names.get(role, role.upper())
        ws = wb.create_sheet(role_name)

        ws.append(["ID", "Category", "Task", "MD"])

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        current_category = None
        category_start_row = 2
        row_num = 2

        for task in wbs_data:
            md = task["effort"].get(role, {}).get("md", 0) or 0
            if md <= 0:
                continue

            category = task["category"]

            if current_category != category:
                if current_category is not None:
                    _add_role_sheet_subtotal(ws, row_num, category_start_row, current_category)
                    row_num += 1

                current_category = category
                category_start_row = row_num

            ws.append([task["id"], category, task["name"], md])
            row_num += 1

        if current_category is not None:
            _add_role_sheet_subtotal(ws, row_num, category_start_row, current_category)

        auto_adjust_columns(ws)


def _add_role_sheet_subtotal(ws, row_num: int, start_row: int, category: str) -> None:
    """Add subtotal row in role sheet."""
    formula = f"=SUM(D{start_row}:D{row_num - 1})"
    ws.append(["", f"{category} Subtotal", "", formula])

    for cell in ws[row_num]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")


def auto_adjust_columns(ws, max_width: int = 60) -> None:
    """Auto-adjust column widths."""
    for col in ws.columns:
        column = col[0].column_letter
        max_length = 0

        for cell in col:
            try:
                cell_value = str(cell.value) if cell.value else ""
                if not cell_value.startswith("="):
                    max_length = max(max_length, len(cell_value))
            except Exception:
                pass

        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column].width = adjusted_width
