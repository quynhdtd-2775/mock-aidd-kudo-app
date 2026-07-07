"""Excel parsing using Pandas and OpenPyXL."""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    import pandas as pd

try:
    import pandas as pd
except ImportError:
    pd = None

from ..parser_base_classes_and_data_models import (
    BaseParser,
    ParsedDocument,
    ParseFailedError,
    Section,
    Table,
)


class ExcelParser(BaseParser):
    """
    Excel parser using Pandas for data extraction.

    Supports: .xlsx, .xls formats
    Features: Multi-sheet handling, table extraction, metadata, hidden sheet detection
    """

    def __init__(
        self,
        max_rows: int | None = None,
        max_cols: int | None = None,
        sheets: list[str] | None = None,
    ):
        """
        Initialize Excel parser.

        Args:
            max_rows: Maximum rows to read per sheet (None = all)
            max_cols: Maximum columns to read per sheet (None = all)
            sheets: Specific sheet names to parse (None = all visible sheets)
        """
        if pd is None:
            raise ImportError("Pandas not installed. Run: pip install pandas openpyxl")
        self.max_rows = max_rows
        self.max_cols = max_cols
        self.sheets = sheets

    def parse(self, source: str) -> ParsedDocument:
        """
        Parse Excel file and extract content from all sheets.

        Args:
            source: Path to Excel file

        Returns:
            ParsedDocument with sheets as sections and tables
        """
        path = Path(source)
        if not path.exists():
            raise ParseFailedError(f"File not found: {source}", source)

        try:
            xlsx = pd.ExcelFile(source)
        except Exception as e:
            raise ParseFailedError(f"Failed to open Excel file: {e}", source, e)

        hidden_sheets = self._detect_hidden_sheets(source)
        visible_sheets = [s for s in xlsx.sheet_names if s not in hidden_sheets]

        if self.sheets:
            target_sheets = [s for s in self.sheets if s in xlsx.sheet_names]
        else:
            target_sheets = visible_sheets

        sections = []
        tables = []
        content_parts = []
        warnings = []
        total_rows = 0
        total_cols = 0

        if hidden_sheets:
            skipped = [s for s in hidden_sheets if s not in target_sheets]
            if skipped:
                warnings.append(f"Hidden sheets detected (skipped): {', '.join(skipped)}")

        for sheet_name in target_sheets:
            try:
                df = self._read_sheet(xlsx, sheet_name)

                if df.empty:
                    warnings.append(f"Sheet '{sheet_name}' is empty")
                    continue

                total_rows += len(df)
                total_cols = max(total_cols, len(df.columns))

                # Create section for this sheet
                section = self._create_section(sheet_name, df)
                sections.append(section)

                # Create table for this sheet
                table = self._create_table(sheet_name, df)
                tables.append(table)

                content_parts.append(f"## {sheet_name}\n\n{table.to_markdown()}")

            except Exception as e:
                warnings.append(f"Failed to parse sheet '{sheet_name}': {e}")

        xlsx.close()

        # Build metadata
        metadata = {
            "sheets": list(target_sheets),
            "sheet_count": len(target_sheets),
            "total_rows": total_rows,
            "total_columns": total_cols,
            "file_size": path.stat().st_size,
            "all_sheets": xlsx.sheet_names,
            "hidden_sheets": hidden_sheets,
        }

        content = "\n\n".join(content_parts)
        title = path.stem

        return ParsedDocument(
            source_path=source,
            source_type="excel",
            title=title,
            content=content,
            sections=sections,
            tables=tables,
            metadata=metadata,
            parse_warnings=warnings,
        )

    def _detect_hidden_sheets(self, source: str) -> list[str]:
        """Detect hidden/very-hidden sheets using openpyxl."""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(source, read_only=True, data_only=True)
            hidden = [
                name for name in wb.sheetnames if wb[name].sheet_state in ("hidden", "veryHidden")
            ]
            wb.close()
            return hidden
        except Exception:
            return []

    def _read_sheet(self, xlsx: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
        """Read a single sheet with optional limits."""
        df = pd.read_excel(
            xlsx,
            sheet_name=sheet_name,
            nrows=self.max_rows,
        )

        # Apply column limit if specified
        if self.max_cols and len(df.columns) > self.max_cols:
            df = df.iloc[:, : self.max_cols]

        # Clean column names
        df.columns = [str(c).strip() for c in df.columns]

        # Fill NaN with empty string for text output
        df = df.fillna("")

        return df

    def _create_section(self, sheet_name: str, df: pd.DataFrame) -> Section:
        """Create a Section from a DataFrame."""
        # Generate content summary
        content_lines = [
            f"Rows: {len(df)}, Columns: {len(df.columns)}",
            f"Columns: {', '.join(df.columns[:10])}",
        ]
        if len(df.columns) > 10:
            content_lines[-1] += f" ... (+{len(df.columns) - 10} more)"

        return Section(
            level=2,
            title=sheet_name,
            content="\n".join(content_lines),
        )

    def _create_table(self, sheet_name: str, df: pd.DataFrame) -> Table:
        """Create a Table from a DataFrame."""
        headers = df.columns.tolist()
        rows = df.astype(str).values.tolist()

        return Table(
            headers=headers,
            rows=rows,
            caption=sheet_name,
        )

    def supports(self, source: str) -> bool:
        """Check if this parser supports the given source."""
        lower = source.lower()
        return lower.endswith(".xlsx") or lower.endswith(".xls")

    def get_sheet_names(self, source: str) -> list[str]:
        """Get list of sheet names without full parsing."""
        try:
            xlsx = pd.ExcelFile(source)
            names = xlsx.sheet_names
            xlsx.close()
            return names
        except Exception:
            return []

    def parse_sheet(self, source: str, sheet_name: str) -> Table:
        """Parse a single sheet and return as Table."""
        try:
            df = pd.read_excel(source, sheet_name=sheet_name)
            df = df.fillna("")
            return Table(
                headers=df.columns.tolist(),
                rows=df.astype(str).values.tolist(),
                caption=sheet_name,
            )
        except Exception as e:
            raise ParseFailedError(f"Failed to parse sheet '{sheet_name}': {e}", source, e)
