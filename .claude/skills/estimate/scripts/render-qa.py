#!/usr/bin/env python3
"""Render discovery Q&A JSON to client-facing markdown and Excel.

Usage:
    python3 scripts/render-qa.py discovery-qa.json -o input/acme/discovery/
    python3 scripts/render-qa.py discovery-qa.json -o input/acme/discovery/ -f md,xlsx
    python3 scripts/render-qa.py discovery-qa.json --lang ja
    cat discovery-qa.json | python3 scripts/render-qa.py - -o output/
"""

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

SCHEMA_PATH = PROJECT_ROOT / "schemas" / "discovery-qa-schema.json"

LABELS = {
    "en": {
        "title": "Discovery Q&A",
        "feature": "Feature",
        "category": "Category",
        "question": "Question",
        "context": "Context",
        "options": "Options",
        "answer": "Answer",
        "priority": "Priority",
        "priority_reason": "Reason",
        "source_files": "Source Documents",
        "date": "Date",
        "total_questions": "Total Questions",
        "answer_placeholder": "_______________",
    },
    "vi": {
        "title": "Khảo sát & Câu hỏi",
        "feature": "Tính năng",
        "category": "Phân loại",
        "question": "Câu hỏi",
        "context": "Bối cảnh",
        "options": "Lựa chọn",
        "answer": "Trả lời",
        "priority": "Độ ưu tiên",
        "priority_reason": "Lý do",
        "source_files": "Tài liệu nguồn",
        "date": "Ngày",
        "total_questions": "Tổng số câu hỏi",
        "answer_placeholder": "_______________",
    },
    "ja": {
        "title": "ディスカバリー Q&A",
        "feature": "機能",
        "category": "カテゴリ",
        "question": "質問",
        "context": "背景",
        "options": "選択肢",
        "answer": "回答",
        "priority": "優先度",
        "priority_reason": "理由",
        "source_files": "参照資料",
        "date": "日付",
        "total_questions": "質問総数",
        "answer_placeholder": "_______________",
    },
}

CATEGORY_DISPLAY = {
    "functional": {"en": "Functional", "vi": "Chức năng", "ja": "機能要件"},
    "technical": {"en": "Technical", "vi": "Kỹ thuật", "ja": "技術要件"},
    "scope": {"en": "Scope", "vi": "Phạm vi", "ja": "スコープ"},
}

PRIORITY_DISPLAY = {
    "critical": {"en": "Critical", "vi": "Quan trọng", "ja": "最重要"},
    "high": {"en": "High", "vi": "Cao", "ja": "高"},
    "medium": {"en": "Medium", "vi": "Trung bình", "ja": "中"},
}


def _display_name(mapping: dict[str, dict[str, str]], key: str, lang: str) -> str:
    return mapping.get(key, {}).get(lang, key)


def validate_data(data: dict) -> None:
    """Validate Q&A data against JSON schema."""
    import jsonschema

    with open(SCHEMA_PATH, encoding="utf-8") as f:
        schema = json.load(f)
    jsonschema.validate(instance=data, schema=schema)


def render_markdown(data: dict, lang: str) -> str:
    """Render Q&A data to grouped markdown."""
    if lang not in LABELS:
        lang = "en"
    lb = LABELS[lang]
    lines: list[str] = []

    lines.append(f"# {lb['title']}: {data['project_name']}")
    lines.append("")
    if data.get("generated_date"):
        lines.append(f"**{lb['date']}:** {data['generated_date']}")
    if data.get("source_files"):
        files_str = ", ".join(data["source_files"])
        lines.append(f"**{lb['source_files']}:** {files_str}")
    lines.append(f"**{lb['total_questions']}:** {len(data.get('questions', []))}")
    lines.append("")
    lines.append("---")
    lines.append("")

    grouped: dict[str, list[dict]] = defaultdict(list)
    for q in data.get("questions", []):
        grouped[q["feature"]].append(q)

    for feature, questions in grouped.items():
        lines.append(f"## {feature}")
        lines.append("")

        for q in questions:
            cat_display = _display_name(CATEGORY_DISPLAY, q["category"], lang)
            pri_display = _display_name(PRIORITY_DISPLAY, q["priority"], lang)

            lines.append(f"### {q['id']}: [{cat_display}] {q['question']}")

            if q.get("context"):
                lines.append(f"**{lb['context']}:** {q['context']}")

            if q.get("options"):
                opts = " / ".join(q["options"])
                lines.append(f"**{lb['options']}:** {opts}")

            answer_val = q.get("answer", "") or lb["answer_placeholder"]
            lines.append(f"**{lb['answer']}:** {answer_val}")

            pri_line = f"**{lb['priority']}:** {pri_display}"
            if q.get("priority_reason"):
                pri_line += f" — {q['priority_reason']}"
            lines.append(pri_line)

            lines.append("")

    return "\n".join(lines)


def render_excel(data: dict, output_path: Path, lang: str) -> None:
    """Render Q&A data to styled Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    if lang not in LABELS:
        lang = "en"
    lb = LABELS[lang]
    wb = Workbook()
    ws = wb.active
    ws.title = lb["title"]

    headers = [
        "#",
        lb["feature"],
        lb["category"],
        lb["question"],
        lb["context"],
        lb["options"],
        lb["answer"],
        lb["priority"],
        lb["priority_reason"],
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    priority_fills = {
        "critical": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "high": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        "medium": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
    }

    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for row_offset, q in enumerate(data.get("questions", []), start=2):
        cat_display = _display_name(CATEGORY_DISPLAY, q["category"], lang)
        pri_display = _display_name(PRIORITY_DISPLAY, q["priority"], lang)
        opts_str = "\n".join(q["options"]) if q.get("options") else ""

        ws.append(
            [
                q["id"],
                q["feature"],
                cat_display,
                q["question"],
                q.get("context", ""),
                opts_str,
                q.get("answer", ""),
                pri_display,
                q.get("priority_reason", ""),
            ]
        )

        if q["priority"] in priority_fills:
            ws.cell(row=row_offset, column=8).fill = priority_fills[q["priority"]]
        for col_idx in range(1, 10):
            ws.cell(row=row_offset, column=col_idx).alignment = wrap_alignment

    col_widths = {
        1: 6,  # #
        2: 20,  # Feature
        3: 14,  # Category
        4: 45,  # Question
        5: 40,  # Context
        6: 30,  # Options
        7: 30,  # Answer
        8: 12,  # Priority
        9: 30,  # Priority Reason
    }
    for col_num, width in col_widths.items():
        col_letter = ws.cell(row=1, column=col_num).column_letter
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Render discovery Q&A JSON to markdown and/or Excel"
    )
    parser.add_argument("input", help="JSON file path or '-' for stdin")
    parser.add_argument(
        "-o",
        "--output-dir",
        type=Path,
        default=None,
        help="Output directory (default: same as input file)",
    )
    parser.add_argument(
        "-f",
        "--formats",
        default="md,xlsx",
        help="Comma-separated: md, xlsx (default: md,xlsx)",
    )
    parser.add_argument(
        "--lang",
        choices=["en", "vi", "ja"],
        default=None,
        help="Output language for headers/labels (default: from JSON or en)",
    )
    args = parser.parse_args()

    try:
        if args.input == "-":
            data = json.load(sys.stdin)
        else:
            with open(args.input, encoding="utf-8") as f:
                data = json.load(f)
    except json.JSONDecodeError as e:
        print(f"Invalid JSON: {e}", file=sys.stderr)
        sys.exit(1)
    except FileNotFoundError:
        print(f"File not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    try:
        validate_data(data)
    except Exception as e:
        print(f"Schema validation failed: {e}", file=sys.stderr)
        sys.exit(1)

    lang = args.lang or data.get("language", "en")
    if lang not in LABELS:
        lang = "en"

    formats = {f.strip().lower() for f in args.formats.split(",")}
    output_dir = args.output_dir or (Path(args.input).parent if args.input != "-" else Path("."))
    output_dir.mkdir(parents=True, exist_ok=True)

    results: list[dict] = []

    if "md" in formats:
        md_text = render_markdown(data, lang)
        md_path = output_dir / "client-qa.md"
        md_path.write_text(md_text, encoding="utf-8")
        results.append({"format": "md", "path": str(md_path)})

    if "xlsx" in formats:
        try:
            xlsx_path = output_dir / "client-qa.xlsx"
            render_excel(data, xlsx_path, lang)
            results.append({"format": "xlsx", "path": str(xlsx_path)})
        except ImportError:
            results.append({"format": "xlsx", "skipped": "openpyxl not installed"})

    print(json.dumps(results, indent=2))


if __name__ == "__main__":
    main()
