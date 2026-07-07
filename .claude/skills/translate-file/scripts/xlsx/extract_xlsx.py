#!/usr/bin/env python3
"""
extract_xlsx.py - Extract translatable cells from XLSX to Markdown chunks.

Each translatable cell becomes a [CELL:SheetName!A1] marker line so
build_xlsx.py can map translations back to the exact cell address.
Formulas, numbers, and empty cells are skipped.
"""

import argparse
import json
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from manifest import create_manifest


def _cell_addr(sheet_name, row, col):
    return f"{sheet_name}!{get_column_letter(col)}{row}"


def _is_translatable(value):
    """Return True for non-empty, non-formula string values."""
    if value is None:
        return False
    if isinstance(value, (int, float, bool)):
        return False
    if isinstance(value, str):
        s = value.strip()
        return bool(s) and not s.startswith("=")
    return False


def extract_xlsx_to_chunks(xlsx_path, temp_dir, chunk_size=6000):
    """Extract translatable cells from xlsx to chunk*.md + xlsx_structure.json.

    Returns number of chunk files created, or 0 on failure.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)

    items = []      # list of (md_line, cell_meta) for chunking
    all_cells = []  # metadata for xlsx_structure.json

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if not _is_translatable(cell.value):
                    continue
                addr = _cell_addr(sheet_name, cell.row, cell.column)
                text = str(cell.value).strip()
                md_line = f"[CELL:{addr}] {text}"
                meta = {
                    "addr": addr,
                    "sheet": sheet_name,
                    "row": cell.row,
                    "col": cell.column,
                    "original": text,
                }
                items.append((md_line, meta))
                all_cells.append(meta)

    if not items:
        print("No translatable text cells found in xlsx")
        return 0

    structure = {
        "schema_version": 1,
        "source_xlsx": os.path.abspath(xlsx_path),
        "cells": all_cells,
    }
    struct_path = os.path.join(temp_dir, "xlsx_structure.json")
    with open(struct_path, "w", encoding="utf-8") as f:
        json.dump(structure, f, ensure_ascii=False, indent=2)
    print(f"Wrote xlsx_structure.json ({len(all_cells)} translatable cells)")

    # Group into chunks
    chunks, current, size = [], [], 0
    for md_line, meta in items:
        block_size = len(md_line) + 2
        if size + block_size > chunk_size and current:
            chunks.append(current)
            current, size = [], 0
        current.append((md_line, meta))
        size += block_size
    if current:
        chunks.append(current)

    # Write input.md (merged for manifest compatibility)
    all_lines = [md for md, _ in items]
    input_md = os.path.join(temp_dir, "input.md")
    with open(input_md, "w", encoding="utf-8") as f:
        f.write("\n\n".join(all_lines))

    chunk_files = []
    for i, group in enumerate(chunks, 1):
        fname = f"chunk{i:04d}.md"
        lines = [md for md, _ in group]
        with open(os.path.join(temp_dir, fname), "w", encoding="utf-8") as f:
            f.write("\n\n".join(lines))
        chunk_files.append(fname)

    create_manifest(temp_dir, chunk_files, input_md)
    print(f"XLSX extracted: {len(chunk_files)} chunks, {len(all_cells)} translatable cells")
    return len(chunk_files)


def main():
    parser = argparse.ArgumentParser(description="Extract XLSX to Markdown chunks")
    parser.add_argument("xlsx_file", help="Input XLSX file")
    parser.add_argument("--temp-dir", required=True, help="Output temp directory")
    parser.add_argument("--chunk-size", type=int, default=6000)
    args = parser.parse_args()

    if not os.path.exists(args.xlsx_file):
        print(f"Error: {args.xlsx_file} not found")
        sys.exit(1)

    os.makedirs(args.temp_dir, exist_ok=True)
    count = extract_xlsx_to_chunks(args.xlsx_file, args.temp_dir, args.chunk_size)
    if count == 0:
        sys.exit(1)
    print(f"Done. Chunks in: {args.temp_dir}")


if __name__ == "__main__":
    main()
