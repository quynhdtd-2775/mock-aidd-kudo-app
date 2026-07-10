#!/usr/bin/env python3
"""
build_xlsx.py - Rebuild XLSX from translated chunks + xlsx_structure.json.

Reads [CELL:SheetName!A1] markers from output_chunk*.md and writes translated
text back into the original xlsx, preserving formulas and all cell formatting.
"""

import argparse
import glob
import json
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


_CELL_MARKER_RE = re.compile(r'^\[CELL:([^\]]+)\]\s*')


def _parse_addr(addr):
    """Parse 'Sheet1!A1' → (sheet_name, row_int, col_int) or (None, None, None)."""
    if "!" not in addr:
        return None, None, None
    sheet, cell_ref = addr.split("!", 1)
    m = re.match(r'^([A-Z]+)(\d+)$', cell_ref)
    if not m:
        return None, None, None
    return sheet, int(m.group(2)), column_index_from_string(m.group(1))


def load_source_xlsx(temp_dir):
    """Read xlsx_structure.json and return source_xlsx path."""
    path = os.path.join(temp_dir, "xlsx_structure.json")
    if not os.path.exists(path):
        print(f"Error: xlsx_structure.json not found in {temp_dir}")
        return None
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("source_xlsx")


def parse_translations(temp_dir):
    """Collect {addr: translated_text} from all output_chunk*.md files."""
    translations = {}
    for chunk_path in sorted(glob.glob(os.path.join(temp_dir, "output_chunk*.md"))):
        with open(chunk_path, "r", encoding="utf-8") as f:
            content = f.read()
        for line in content.splitlines():
            m = _CELL_MARKER_RE.match(line.strip())
            if not m:
                continue
            addr = m.group(1)
            text = line.strip()[m.end():].strip()
            translations[addr] = text
    return translations


def build_xlsx(temp_dir, output_path=None):
    """Write translations back into xlsx. Returns True on success."""
    source_xlsx = load_source_xlsx(temp_dir)
    if not source_xlsx or not os.path.exists(source_xlsx):
        print(f"Error: source xlsx not found: {source_xlsx}")
        return False

    translations = parse_translations(temp_dir)
    if not translations:
        print("Error: no [CELL:...] translations found in output_chunk*.md files")
        return False

    wb = openpyxl.load_workbook(source_xlsx)
    updated = 0
    missing = []

    for addr, translated_text in translations.items():
        sheet_name, row, col = _parse_addr(addr)
        if sheet_name is None:
            continue
        if sheet_name not in wb.sheetnames:
            missing.append(addr)
            continue
        wb[sheet_name].cell(row=row, column=col).value = translated_text
        updated += 1

    if missing:
        print(f"WARNING: {len(missing)} addresses not found in workbook: {missing[:5]}")

    if output_path is None:
        output_path = os.path.join(temp_dir, "output.xlsx")

    wb.save(output_path)
    size = os.path.getsize(output_path)
    print(f"XLSX rebuilt: {updated} cells translated → {output_path} ({size:,} bytes)")
    return True


def main():
    parser = argparse.ArgumentParser(description="Rebuild XLSX from translated chunks")
    parser.add_argument("--temp-dir", required=True, help="Temp directory containing output_chunk*.md")
    parser.add_argument("--output", default=None, help="Output xlsx path (default: <temp-dir>/output.xlsx)")
    args = parser.parse_args()

    if not os.path.isdir(args.temp_dir):
        print(f"Error: temp dir not found: {args.temp_dir}")
        sys.exit(1)

    if not build_xlsx(args.temp_dir, args.output):
        sys.exit(1)


if __name__ == "__main__":
    main()
