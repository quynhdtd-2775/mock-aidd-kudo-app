"""Style fingerprint for xlsx fidelity verification.

Captures per-cell style + sheet-level layout so two worksheets can be compared
on STYLE alone (font/size/fill/align/border/number-format/merges/widths/heights/
data-validations), independent of cell values. Used by verify_format.py.

Reused by build_api_design.py for self-checks. No third-party deps beyond openpyxl.
"""
from openpyxl.utils import get_column_letter


def _color(c):
    """Normalize an openpyxl Color to a comparable list [type, value, tint]."""
    if c is None:
        return None
    t = getattr(c, "type", None)
    tint = round(float(getattr(c, "tint", 0) or 0), 4)
    if t == "rgb":
        return ["rgb", c.rgb, tint]
    if t == "theme":
        return ["theme", c.theme, tint]
    if t == "indexed":
        return ["indexed", c.indexed, tint]
    return [str(t), getattr(c, "value", None), tint]


def _side(s):
    if s is None or s.style is None:
        return None
    return [s.style, _color(s.color)]


def cell_style(cell):
    """Per-cell style dict (no value)."""
    f, fill, al, b = cell.font, cell.fill, cell.alignment, cell.border
    has_fill = bool(fill and fill.patternType)
    return {
        "font": [f.name, f.size, bool(f.bold), bool(f.italic),
                 (f.underline or None), _color(f.color)],
        "fill": [fill.patternType if fill else None,
                 _color(fill.fgColor) if has_fill else None],
        "align": [al.horizontal, al.vertical, bool(al.wrap_text)],
        "border": [_side(b.left), _side(b.right), _side(b.top), _side(b.bottom)],
        "numfmt": cell.number_format,
    }


def sheet_fingerprint(ws, include_values=False):
    """Fingerprint a worksheet: every cell's style + layout metadata."""
    cells = {}
    for row in ws.iter_rows():
        for cell in row:
            entry = cell_style(cell)
            if include_values:
                entry["value"] = None if cell.value is None else str(cell.value)
            cells[cell.coordinate] = entry
    merged = sorted(str(m) for m in ws.merged_cells.ranges)
    cols = {k: round(v.width, 3) for k, v in ws.column_dimensions.items()
            if v.width is not None}
    rows = {str(k): round(v.height, 3) for k, v in ws.row_dimensions.items()
            if v.height is not None}
    dvs = []
    for dv in ws.data_validations.dataValidation:
        dvs.append({"type": dv.type, "formula1": dv.formula1, "sqref": str(dv.sqref)})
    dvs.sort(key=lambda d: (d.get("sqref") or ""))
    return {
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "merged": merged,
        "col_widths": cols,
        "row_heights": rows,
        "data_validations": dvs,
        "cells": cells,
    }


def col_of(coord):
    """Column letters from a coord like 'B12' -> 'B'."""
    return "".join(ch for ch in coord if ch.isalpha())


def row_of(coord):
    return int("".join(ch for ch in coord if ch.isdigit()))
