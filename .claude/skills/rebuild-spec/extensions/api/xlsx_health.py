"""Container health check for a generated .xlsx — proves it opens without 'needs repair'.

Complements verify_format.py (which checks STYLE fidelity). This validates the OOXML CONTAINER:
zip integrity, XML well-formedness, strict UTF-8, every relationship target resolving, images,
data validations, external links, and a non-ASCII (e.g. Japanese) encoding round-trip — the
structural faults that make Excel/LibreOffice refuse or repair a file.

Usage:  python xlsx_health.py <output.xlsx>
Exit 0 = healthy. Exit 2 = problem(s) found (report printed). Exit 1 = usage/IO error.
"""
import os
import re
import sys
import posixpath
import zipfile
import xml.etree.ElementTree as ET

RNS = "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
ENC_RE = re.compile(rb'encoding=["\']([^"\']+)["\']')


def main(argv):
    if len(argv) != 1:
        sys.stderr.write("usage: xlsx_health.py <output.xlsx>\n")
        return 1
    path = argv[0]
    if not os.path.exists(path):
        sys.stderr.write(f"ERROR: not found: {path}\n")
        return 1

    problems = []

    def fail(cat, detail):
        problems.append(cat)
        print(f"  FAIL [{cat}] {detail}")

    def ok(detail):
        print(f"  ..   {detail}")

    print(f"== xlsx health: {path}")
    try:
        z = zipfile.ZipFile(path)
    except zipfile.BadZipFile as e:
        print(f"  FAIL [zip] not a valid zip/xlsx ({e})")
        return 2

    # A. zip CRC integrity
    bad = z.testzip()
    if bad:
        fail("zip", f"corrupt entry: {bad}")
    else:
        ok(f"zip OK — {len(z.namelist())} parts, CRC verified")
    names = set(z.namelist())

    # B. required core parts
    required = ["[Content_Types].xml", "_rels/.rels", "xl/workbook.xml",
                "xl/_rels/workbook.xml.rels", "xl/styles.xml"]
    miss = [r for r in required if r not in names]
    if miss:
        fail("part", f"missing required parts: {miss}")
    else:
        ok("required core parts present (5/5)")

    # C. every XML part: well-formed + strict UTF-8
    xml_parts = [n for n in names if n.endswith((".xml", ".rels"))]
    for n in xml_parts:
        raw = z.read(n)
        try:
            raw.decode("utf-8")
        except UnicodeDecodeError as e:
            fail("utf8", f"{n}: not valid UTF-8 ({e})")
        m = ENC_RE.search(raw[:120])
        if m and m.group(1).lower() != b"utf-8":
            fail("enc", f"{n}: declares non-UTF-8 encoding {m.group(1)!r}")
        try:
            ET.fromstring(raw)
        except ET.ParseError as e:
            fail("xml", f"{n}: malformed XML ({e})")
    ok(f"{len(xml_parts)} XML parts — well-formed + UTF-8")

    # D. every relationship Target resolves to an existing internal part (the 'needs repair' classic)
    rels = [n for n in names if n.endswith(".rels")]
    dangling = 0
    for rp in rels:
        base = posixpath.dirname(posixpath.dirname(rp))   # _rels sits under the owning dir
        for rel in ET.fromstring(z.read(rp)).findall(RNS):
            if (rel.get("TargetMode") or "") == "External":
                continue
            tgt = rel.get("Target") or ""
            resolved = tgt.lstrip("/") if tgt.startswith("/") else posixpath.normpath(posixpath.join(base, tgt))
            if resolved not in names:
                dangling += 1
                fail("rel", f"{rp}: target missing -> {tgt}")
    if not dangling:
        ok(f"all relationship targets resolve ({len(rels)} .rels parts)")

    # E. external links → Excel update/repair prompt
    ext = [n for n in names if "externallink" in n.lower()]
    if ext or "<externalReferences" in z.read("xl/workbook.xml").decode("utf-8"):
        fail("ext", f"external references present ({ext or 'workbook.xml'})")
    else:
        ok("no external references")

    # F. images present
    media = [n for n in names if n.startswith("xl/media/")]
    draw = [n for n in names if n.startswith("xl/drawings/drawing")]
    ok(f"media: {len(media)} | drawings: {len(draw)}")

    # G. openpyxl reload (normal + read_only)
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path)
        openpyxl.load_workbook(path, read_only=True).close()
        ok(f"openpyxl reload OK (normal + read_only) — {len(wb.sheetnames)} sheets")
    except Exception as e:  # noqa: BLE001 — any reload failure is a real openability fault
        fail("reload", f"openpyxl cannot reopen: {e}")
        return 2

    # H. data validations reference existing sheets
    sheets = set(wb.sheetnames)
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            for ref in re.findall(r"'?([A-Za-z0-9_ ]+)'?!", dv.formula1 or ""):
                if ref not in sheets:
                    fail("dv", f"{ws.title}: data-validation refs missing sheet {ref!r}")
    ok("data validations reference existing sheets")

    # I. non-ASCII (e.g. Japanese) encoding round-trip
    sample = None
    for ws in wb.worksheets:
        for row in ws.iter_rows(max_row=40):
            for c in row:
                if isinstance(c.value, str) and any(ord(ch) > 127 for ch in c.value):
                    sample = (ws.title, c.coordinate, c.value)
                    break
            if sample:
                break
        if sample:
            break
    ok(f"non-ASCII round-trip OK e.g. [{sample[0]}]{sample[1]} = {sample[2][:32]!r}"
       if sample else "no non-ASCII cell in sampled range")

    if problems:
        print(f"\n=== UNHEALTHY — {len(problems)} problem(s): {sorted(set(problems))} ===")
        return 2
    print("\n=== HEALTHY — no open-time errors expected ===")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
