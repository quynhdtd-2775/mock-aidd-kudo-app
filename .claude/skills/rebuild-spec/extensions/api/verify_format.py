"""Fidelity verifier — proves the generated workbook is style-identical to the sample.

Golden = the Sun* sample. Two modes:
  --make-golden SAMPLE GOLDEN_JSON   capture baseline (BEFORE build)
  GOLDEN_JSON OUTPUT_XLSX            verify output has 0 style drift (AFTER build)

Fidelity model:
  * Detail sheets (anything not structural/template) are compared to the
    API_Detail_Template fingerprint, IGNORING cell values. Repeatable bands
    (param rows 13+, response rows 27+) let rows grow while their style must
    still match the band's pattern row — cell style AND single-row merges.
  * Structural sheets compared to their own golden over shared cells; grown
    body rows compared to the sheet's pattern row (style only).
  * Data validations: the set of (type, formula1) in output must cover golden's
    (catches copy_worksheet dropping dropdowns).
  * Content completeness: reads api-content.json beside the output and FAILS if any op
    exceeds the template's fixed area (params>10 or responses>7) — i.e. the builder dropped
    data. "SEALED" therefore means format-perfect AND content-complete, not format alone.
    Skipped (not failed) when no api-content.json is present (standalone xlsx).
  * Coverage: every extracted operation must map to exactly one detail sheet (n_ops == n_detail);
    catches a sheet silently dropped or added between extract and build. Skipped without api-content.json.

Exit 0 = SEALED (0 drift + content-complete + full coverage). Exit 2 = drift, dangling-ref, missing
logo, truncation, or coverage mismatch (report printed). Exit 1 = usage.
"""
import os
import sys
import re
import json
import argparse
try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl is required — install with `pip install openpyxl`, "
             "or run via the kit venv: .claude/skills/.venv/bin/python3")
from style_fingerprint import sheet_fingerprint, col_of, row_of

# Template's fixed per-detail-sheet capacity — mirrors build_api_design.py (params[:10],
# responses[:7]). An op exceeding either had data SILENTLY DROPPED by the builder, so the
# workbook is style-perfect but content-incomplete. The completeness gate below fails on it.
CAP_PARAMS, CAP_RESP = 10, 7

STRUCTURAL = ["History of changes", "Common Conventions", "Status code", "API List", "Appendix"]
TEMPLATE = "API_Detail_Template"
SAMPLE_DETAIL = "SamplePOST apiv1login"   # the worked example — verify like a detail sheet

# Repeatable data bands: a cell at row >= start is compared to pattern_row's style
# (takes precedence over exact golden match — body rows must follow the data pattern,
# not the sample's specific rows, which include a trailing 'cap' row we overwrite).
BANDS = {
    "__detail__": [{"start": 30, "pattern": 28}],   # rows 27-29 = template (exact); 30+ grow from middle row 28
    "API List": [{"start": 8, "pattern": 11}],       # body rows 8.. follow data row 11
    "Status code": [{"start": 11, "pattern": 11}],   # body rows 11.. follow data row 11
}
STYLE_ATTRS = ("font", "fill", "align", "border", "numfmt")
SHEET_REF = re.compile(r"'([^']+)'!|(?<![A-Za-z0-9_])([A-Za-z_][A-Za-z0-9_ ]*)!")


def dangling_refs(wb):
    """Formulas referencing a sheet that does not exist → Excel 'needs repair'."""
    names = set(wb.sheetnames)
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    for m in SHEET_REF.finditer(c.value):
                        ref = m.group(1) or m.group(2)
                        if ref and ref not in names:
                            out.append((ws.title, c.coordinate, ref))
    return out


def _band_pattern_row(bands, r):
    best = None
    for b in bands:
        if b["start"] <= r and (best is None or b["start"] > best["start"]):
            best = b
    return best["pattern"] if best else None


def _cmp_cell(ref_style, out_style, sheet, coord, drift):
    for attr in STYLE_ATTRS:
        if ref_style.get(attr) != out_style.get(attr):
            drift.append((sheet, coord, attr,
                          json.dumps(ref_style.get(attr), ensure_ascii=False),
                          json.dumps(out_style.get(attr), ensure_ascii=False)))


def diff_sheet(ref_fp, out_fp, sheet, bands, strict_merge_extra=True):
    """Style-only diff of out_fp against ref_fp, band-aware for grown rows.

    strict_merge_extra=False (structural sheets) tolerates extra merges from
    growing the body (e.g. API List category groups merged vertically).
    """
    drift = []
    ref_cells, out_cells = ref_fp["cells"], out_fp["cells"]
    for coord, out_style in out_cells.items():
        pr = _band_pattern_row(bands, row_of(coord))
        if pr is not None:                            # in a data band — match pattern row
            ref = ref_cells.get(f"{col_of(coord)}{pr}")
            if ref is not None:
                _cmp_cell(ref, out_style, sheet, coord, drift)
            else:
                drift.append((sheet, coord, "unexpected-cell", "", "present(no-pattern)"))
        elif coord in ref_cells:                      # fixed region — exact match
            _cmp_cell(ref_cells[coord], out_style, sheet, coord, drift)
        else:
            drift.append((sheet, coord, "unexpected-cell", "", "present(no-golden)"))
    # merges: exact golden merges must be present; extra single-row merges in a
    # band must match the pattern row's merge (shifted).
    ref_m, out_m = set(ref_fp["merged"]), set(out_fp["merged"])
    body_start = min((b["start"] for b in bands), default=None)
    for m in ref_m - out_m:
        if not strict_merge_extra and body_start and row_of(m.split(":")[0]) >= body_start:
            continue  # body merges regrouped (e.g. category spans) — allowed
        drift.append((sheet, m, "merge-missing", m, ""))
    for m in out_m - ref_m:
        a, b = m.split(":")
        r1, r2 = row_of(a), row_of(b)
        pr = _band_pattern_row(bands, r1)
        if pr and r1 == r2:  # single-row band merge — shift to pattern row
            shifted = f"{col_of(a)}{pr}:{col_of(b)}{pr}"
            if shifted in ref_m:
                continue
        if not strict_merge_extra:
            continue  # structural body grew (e.g. category group merge) — allowed
        drift.append((sheet, m, "merge-extra", "", m))
    # data validations: (type, formula1) coverage
    ref_dv = {(d["type"], d["formula1"]) for d in ref_fp["data_validations"]}
    out_dv = {(d["type"], d["formula1"]) for d in out_fp["data_validations"]}
    for d in ref_dv - out_dv:
        drift.append((sheet, "-", "data-validation-missing", str(d), ""))
    return drift


def content_truncations(output_path):
    """Detect ops whose param/response count exceeds the template area — i.e. data the builder
    dropped. Reads api-content.json beside the output xlsx (written there by extract_api_content).

    Returns (checked, truncations, n_ops). checked=False when no intermediate is present (standalone
    xlsx) → the completeness gate is skipped, not failed. truncations = list of
    (sheet_name, n_params, n_responses) for every over-cap op. n_ops = operations extracted (used
    by the coverage gate: one op must map to exactly one detail sheet).
    """
    cpath = os.path.join(os.path.dirname(os.path.abspath(output_path)), "api-content.json")
    if not os.path.exists(cpath):
        return False, [], 0
    try:
        with open(cpath, encoding="utf-8") as f:
            content = json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        print(f"[WARN] api-content.json unreadable ({e}) — completeness check skipped")
        return False, [], 0
    if not isinstance(content, list):
        print("[WARN] api-content.json is not a list — completeness check skipped")
        return False, [], 0
    trunc = []
    for rec in content:
        if not isinstance(rec, dict):
            continue
        np = len(rec.get("params") or [])
        nr = len(rec.get("responses") or [])
        if np > CAP_PARAMS or nr > CAP_RESP:
            trunc.append((rec.get("sheet_name", "?"), np, nr))
    return True, trunc, len(content)


def make_golden(sample_path, golden_path):
    wb = openpyxl.load_workbook(sample_path)
    golden = {"structural": {}, "template_name": TEMPLATE}
    for name in STRUCTURAL:
        if name in wb.sheetnames:
            golden["structural"][name] = sheet_fingerprint(wb[name])
    golden["template"] = sheet_fingerprint(wb[TEMPLATE])
    with open(golden_path, "w") as f:
        json.dump(golden, f, ensure_ascii=False, indent=1)
    print(f"golden written: {golden_path}")
    print(f"  structural: {list(golden['structural'])}")
    print(f"  template:   {TEMPLATE} ({golden['template']['max_row']}x{golden['template']['max_col']})")


def verify(golden_path, output_path):
    with open(golden_path) as f:
        golden = json.load(f)
    wb = openpyxl.load_workbook(output_path)
    tmpl_fp = golden["template"]
    known = set(golden["structural"]) | {golden["template_name"]}
    drift, n_detail = [], 0
    for name in wb.sheetnames:
        out_fp = sheet_fingerprint(wb[name])
        if name in golden["structural"]:
            drift += diff_sheet(golden["structural"][name], out_fp, name,
                                BANDS.get(name, []), strict_merge_extra=False)
        elif name == golden["template_name"]:
            drift += diff_sheet(tmpl_fp, out_fp, name, [])  # template must be untouched
        else:
            n_detail += 1
            drift += diff_sheet(tmpl_fp, out_fp, name, BANDS["__detail__"])
    dangling = dangling_refs(wb)
    no_logo = [ws.title for ws in wb.worksheets if not ws._images]
    checked, trunc, n_ops = content_truncations(output_path)
    trunc_note = f"{len(trunc)}" if checked else "skipped(no api-content.json)"
    cov_note = f"{n_detail}/{n_ops}" if checked else "skipped"
    print(f"sheets: {len(wb.sheetnames)} ({n_detail} detail vs template) | drift: {len(drift)} | "
          f"dangling-refs: {len(dangling)} | sheets-without-logo: {len(no_logo)} | "
          f"truncations: {trunc_note} | coverage(sheets/ops): {cov_note}")
    if dangling:
        print("\n=== DANGLING SHEET REFERENCES (Excel will fail to open) ===")
        for s, c, ref in dangling[:50]:
            print(f"  [{s}] {c} -> missing sheet {ref!r}")
        return 2
    if no_logo:
        print("\n=== SHEETS MISSING LOGO ===")
        print("  " + ", ".join(no_logo[:80]))
        return 2
    if trunc:
        print(f"\n=== CONTENT TRUNCATED ({len(trunc)} ops exceed template area — data dropped) ===")
        print(f"  caps: params<={CAP_PARAMS}, responses<={CAP_RESP}")
        for sn, np, nr in trunc[:80]:
            flags = []
            if np > CAP_PARAMS:
                flags.append(f"params {np}>{CAP_PARAMS}")
            if nr > CAP_RESP:
                flags.append(f"responses {nr}>{CAP_RESP}")
            print(f"  [{sn}] {', '.join(flags)}")
        print("  Style is perfect but the deliverable is INCOMPLETE — not SEALED.")
        return 2
    if checked and n_ops != n_detail:
        print(f"\n=== COVERAGE MISMATCH ({n_ops} operations extracted, {n_detail} detail sheets built) ===")
        print("  Every API operation must map to exactly one detail sheet — one was dropped or added.")
        return 2
    if drift:
        print("\n=== STYLE DRIFT ===")
        print(f"{'sheet':<24} {'cell':<8} {'attr':<22} golden -> output")
        for s, c, a, g, o in drift[:200]:
            print(f"{s[:23]:<24} {c:<8} {a:<22} {g[:40]} -> {o[:40]}")
        if len(drift) > 200:
            print(f"... +{len(drift) - 200} more")
        return 2
    seal = "0 style drift" + ("" if checked else "; content-completeness NOT checked (no api-content.json)")
    print(f"RESULT: SEALED — {seal}. Output is format-identical to sample"
          + (" and content-complete." if checked else "."))
    return 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("a")
    ap.add_argument("b")
    ap.add_argument("--make-golden", action="store_true")
    args = ap.parse_args()
    if args.make_golden:
        make_golden(args.a, args.b)
        return 0
    return verify(args.a, args.b)


if __name__ == "__main__":
    sys.exit(main())
