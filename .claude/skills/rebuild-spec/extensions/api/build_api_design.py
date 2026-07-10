"""Build the API Design workbook by cloning a Sun*-style template (BM-2-901-52).

Project-agnostic (rebuild-spec --api-doc extension). Loads the template (keeping ALL styles),
clones API_Detail_Template per API and writes ONLY cell .value — never touches style. Re-applies
the two data validations AND the logo that copy_worksheet drops. Populates API List / History /
Status code from the JSON produced by extract_api_content.py.

Usage:
  python build_api_design.py --in-dir DIR --out OUT.xlsx
        [--template api-design-template.xlsx] [--logo sun-logo.png]
        [--system-name NAME] [--creator NAME] [--date YYYY-MM-DD]
Defaults: template + logo are the bundled assets next to this script.
"""
import os
import sys
import json
import argparse
from copy import copy
try:
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.drawing.image import Image
except ImportError:
    sys.exit("ERROR: openpyxl is required — install with `pip install openpyxl`, "
             "or run via the kit venv: .claude/skills/.venv/bin/python3")
from style_fingerprint import col_of, row_of

HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_SHEET = "API_Detail_Template"
SAMPLE_DETAIL = "SamplePOST apiv1login"
LOGO_W, LOGO_H = 46, 26


def copy_row_style(ws, src, dst, ncols):
    for c in range(1, ncols + 1):
        ws.cell(dst, c)._style = copy(ws.cell(src, c)._style)
    if ws.row_dimensions[src].height is not None:
        ws.row_dimensions[dst].height = ws.row_dimensions[src].height


def add_dv(ws, rng, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(rng)


def add_logo(ws, logo):
    if ws._images or not logo or not os.path.exists(logo):
        return
    img = Image(logo)
    img.width, img.height = LOGO_W, LOGO_H
    ws.add_image(img, "A1")


def _csafe(v):
    """Neutralize Excel formula injection: a cell value beginning with = + - @ is treated
    as a formula by Excel. Prefix a single quote to force literal-text interpretation for
    any untrusted (artifact/swagger-derived) string. Non-strings pass through unchanged."""
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
        return "'" + v
    return v


def fill_detail(ws, rec, meta):
    ws["E6"], ws["E7"], ws["E8"] = _csafe(rec["api_name"]), _csafe(rec["endpoint"]), _csafe(rec["method"])
    ws["E9"], ws["E10"] = _csafe(rec["description"]), _csafe(rec["media_type"])
    ws["K6"], ws["K7"] = meta["date"], meta["creator"]
    ws["K8"], ws["K9"] = meta["date"], meta["creator"]
    for i, p in enumerate(rec["params"][:10]):
        r = 13 + i
        ws.cell(r, 2, _csafe(p["param_type"])); ws.cell(r, 3, _csafe(p["key"])); ws.cell(r, 5, _csafe(p["data_type"]))
        ws.cell(r, 6, _csafe(p["values"])); ws.cell(r, 7, _csafe(p["description"])); ws.cell(r, 9, _csafe(p["note"]))
    nresp = min(len(rec["responses"]), 7)
    for r in range(30, max(31, 27 + nresp)):          # normalize template's bold row 30; grow as needed
        copy_row_style(ws, 28, r, 13)
        ws.merge_cells(f"D{r}:H{r}"); ws.merge_cells(f"I{r}:L{r}")
    for i, rsp in enumerate(rec["responses"][:7]):
        r = 27 + i
        ws.cell(r, 1, i + 1); ws.cell(r, 2, _csafe(rsp["status"])); ws.cell(r, 3, _csafe(rsp["description"]))
        ws.cell(r, 4, _csafe(rsp["example"])); ws.cell(r, 9, _csafe(rsp["note"]))
    add_dv(ws, "B13:B22", "Appendix!$N$4:$N$8")
    add_dv(ws, "E13:E22", "Appendix!$K$4:$K$8")


def fill_history(ws, meta):
    ws["A8"] = 1; ws["B8"] = "Ver 0.1"; ws["E8"] = "Add new"; ws["F8"] = "All"
    ws["H8"] = f"Initial API design — {meta.get('provenance', 'generated from the OpenAPI spec')}"
    ws["M8"] = meta["date"]; ws["N8"] = meta["creator"]


def fill_status(ws, codes):
    pat = [copy(ws.cell(11, c)._style) for c in range(1, 10)]
    h11 = ws.row_dimensions[11].height
    last = ws.max_row                                 # template sample extent — trailing rows normalized too
    for m in list(ws.merged_cells.ranges):
        if row_of(str(m).split(":")[0]) >= 11:
            ws.unmerge_cells(str(m))
    # one loop over data + any leftover template sample rows: every row in the band gets the pattern
    # style + merges; data rows carry values, trailing rows are blanked (else stale sample = style drift).
    for r in range(11, max(11 + len(codes), last + 1)):
        for c in range(1, 10):
            ws.cell(r, c)._style = copy(pat[c - 1])
        if h11 is not None:
            ws.row_dimensions[r].height = h11
        ws.merge_cells(f"B{r}:C{r}"); ws.merge_cells(f"D{r}:G{r}"); ws.merge_cells(f"H{r}:I{r}")
        i = r - 11
        if i < len(codes):
            ws.cell(r, 1, i + 1); ws.cell(r, 2, codes[i]["code"]); ws.cell(r, 4, codes[i]["description"])
            ws.cell(r, 8).value = None                # clear sample cross-ref via .value (cell(r,8,None) would no-op)
        else:
            for c in (1, 2, 4, 8):                    # blank leftover sample row (anchors only — merged
                ws.cell(r, c).value = None            # non-anchor cells are read-only in openpyxl)


def fill_api_list(ws, rows):
    n = len(rows)
    pat = [copy(ws.cell(11, c)._style) for c in range(1, 12)]
    h11 = ws.row_dimensions[11].height
    last = ws.max_row
    for m in list(ws.merged_cells.ranges):
        if row_of(str(m).split(":")[0]) >= 8:
            ws.unmerge_cells(str(m))
    for r in range(8, max(8 + n, last + 1)):
        for c in range(1, 12):
            ws.cell(r, c)._style = copy(pat[c - 1])
        if h11 is not None:
            ws.row_dimensions[r].height = h11
        ws.merge_cells(f"B{r}:D{r}"); ws.merge_cells(f"F{r}:G{r}")
        ws.merge_cells(f"H{r}:I{r}"); ws.merge_cells(f"J{r}:K{r}")
        i = r - 8
        if i < n:
            ws.cell(r, 1, i + 1); ws.cell(r, 2, rows[i]["category"])
            ws.cell(r, 5, rows[i]["screen_feature"]); ws.cell(r, 6, rows[i]["api_url"])
            ws.cell(r, 8, rows[i]["detail_description"]); ws.cell(r, 10, rows[i]["note"])
        else:
            for c in (1, 2, 5, 6, 8, 10):             # blank leftover sample row (anchors only)
                ws.cell(r, c).value = None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in-dir", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--template", default=os.path.join(HERE, "api-design-template.xlsx"))
    ap.add_argument("--logo", default=os.path.join(HERE, "sun-logo.png"))
    ap.add_argument("--system-name", default="System")
    ap.add_argument("--creator", default="Sun Asterisk Vietnam")
    ap.add_argument("--date", default="")
    ap.add_argument("--provenance", default="generated from the OpenAPI spec",
                    help="source provenance noted in the History sheet (swagger vs derived)")
    args = ap.parse_args()
    meta = {"creator": args.creator, "date": args.date, "provenance": args.provenance}

    content = json.load(open(os.path.join(args.in_dir, "api-content.json"), encoding="utf-8"))
    api_list = sorted(json.load(open(os.path.join(args.in_dir, "api-list.json"), encoding="utf-8")),
                      key=lambda r: (r["category"], r["api_url"]))
    status = json.load(open(os.path.join(args.in_dir, "status-codes.json"), encoding="utf-8"))

    wb = openpyxl.load_workbook(args.template)
    for name in ("History of changes", "Status code", "API List", TEMPLATE_SHEET):
        if name not in wb.sheetnames:
            raise SystemExit(f"TEMPLATE missing expected sheet: {name!r}")
    if SAMPLE_DETAIL in wb.sheetnames:
        del wb[SAMPLE_DETAIL]
    fill_history(wb["History of changes"], meta)
    fill_status(wb["Status code"], status)
    fill_api_list(wb["API List"], api_list)

    template = wb[TEMPLATE_SHEET]
    for rec in content:
        if len(rec["responses"]) > 7:
            print(f"WARNING: {rec['sheet_name']} has {len(rec['responses'])} responses — truncated to 7")
        ws = wb.copy_worksheet(template)
        ws.title = rec["sheet_name"]
        fill_detail(ws, rec, meta)

    for ws in wb.worksheets:
        add_logo(ws, args.logo)

    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    wb.save(args.out)
    print(f"saved: {args.out}")
    print(f"sheets: {len(wb.sheetnames)} | detail: {len(content)} | "
          f"api-list rows: {len(api_list)} | status rows: {len(status)}")


if __name__ == "__main__":
    main()
